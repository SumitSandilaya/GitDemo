import openpyxl


class HomePageData:
    test_HomePageData = [{"Fname": "Sumit", "Email": "sumitsandilaya@gmail.com", "Lname": "Sandilaya", "gender": "Male"}, {"Fname": "Richa", "Email": "richasandilaya@gmail.com", "Lname": "Sandilaya", "gender": "Female"}]

    @staticmethod
    def getTestData(test_case_name):
        workbook = openpyxl.load_workbook(
            "D:\\Python_Selenium\\PythonSelFramework\\ExcelData\\PytestDemo.xlsx")
        sheet = workbook.active
        Dict = {}
        for i in range(1, sheet.max_row + 1):  # to get row
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                #Dictionary
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                    # print(sheet.cell(row=i,column=j).value)

        return [Dict]
