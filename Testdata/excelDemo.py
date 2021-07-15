import openpyxl

workbook = openpyxl.load_workbook("D:\\Python_Selenium\\PythonSelFramework\\ExcelData\\PytestDemo.xlsx")

sheet = workbook.active

Dict = {}
cell = sheet.cell(row=1, column=2)
print(cell.value)

sheet.cell(row=4, column=2).value = "Vivaan"
print(sheet.cell(row=4, column=2).value)

print(sheet.max_row)
print(sheet.max_column)

print(sheet['A5'].value)


for i in range(1, sheet.max_row+1): # to get row
    if sheet.cell(row=i, column=1).value == "TestCase2":
        for j in range(2, sheet.max_column+1):
            #Dictionary
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            #print(sheet.cell(row=i,column=j).value)

print(Dict)